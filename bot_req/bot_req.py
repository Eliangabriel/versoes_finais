import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, NamedStyle, PatternFill
from copy import copy
from datetime import datetime
import os

# Função para abrir a janela de seleção de arquivo
def selecionar_arquivo():
    def abrir_dialogo():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            label_arquivo.config(text=f"Arquivo selecionado: {file_path}")
            btn_iniciar.config(state=tk.NORMAL)  # Ativa o botão "Iniciar Processo"
            btn_iniciar.file_path = file_path  # Armazena o caminho do arquivo no botão

    def iniciar_processo():
        file_path = getattr(btn_iniciar, 'file_path', None)
        if file_path:
            editar_planilha(file_path)
            messagebox.showinfo("Processo", "Processo concluído!")

    # Cria a janela principal
    root = tk.Tk()
    root.title("Selecionar Arquivo")
    root.geometry("600x400")
    
    # Aplicar o tema 'clam'
    style = ttk.Style()
    style.theme_use('clam')

    # Adiciona um frame para centralizar os widgets
    frame = ttk.Frame(root, padding="10")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # Botão para abrir o diálogo de seleção de arquivo
    btn_selecionar = ttk.Button(frame, text="Selecionar Arquivo", command=abrir_dialogo)
    btn_selecionar.grid(row=1, column=0, pady=10)

    # Label para mostrar o arquivo selecionado
    label_arquivo = ttk.Label(frame, text="Nenhum arquivo selecionado")
    label_arquivo.grid(row=2, column=0, pady=10)

    # Botão para iniciar o processo
    btn_iniciar = ttk.Button(frame, text="Iniciar Processo", command=iniciar_processo, state=tk.DISABLED)
    btn_iniciar.grid(row=3, column=0, pady=10)

    # Inicia o loop principal da aplicação
    root.mainloop()

# Função principal para editar a planilha
def editar_planilha(file_path):
    wb_origem = load_workbook(file_path)
    ws_base = wb_origem.active  # Seleciona a aba ativa por padrão

    # Criando um novo arquivo Excel com duas abas
    wb_novo = Workbook()
    ws_novo1 = wb_novo.active
    ws_novo1.title = "Planilha1"
    ws_novo2 = wb_novo.create_sheet(title="Planilha2")

    # Copiando dados e formatação
    for row in ws_base.iter_rows():
        for cell in row:
            for ws in [ws_novo1, ws_novo2]:
                new_cell = ws[cell.coordinate]
                new_cell.value = cell.value
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill) if cell.fill else PatternFill()
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)

    # Excluindo colunas da Planilha1
    for coluna in reversed(['J']):
        ws_novo1.delete_cols(column_index_from_string(coluna))

    # Excluindo colunas da Planilha2
    for coluna in reversed(['G', 'J', 'M', 'N', 'O']):
        ws_novo2.delete_cols(column_index_from_string(coluna))

    # Aplicando filtros
    ws_novo1.auto_filter.ref = ws_novo1.dimensions
    ws_novo2.auto_filter.ref = ws_novo2.dimensions

    # Ajustando tamanho das colunas
    colunas_largura = {'A': 12, 'B': 19, 'C': 19, 'D': 22, 'E': 11, 'F': 30,
                        'G': 30, 'H': 23, 'I': 20, 'J': 12, 'K': 12, 'L': 21,
                        'M': 25, 'N': 47}
    for col, largura in colunas_largura.items():
        ws_novo1.column_dimensions[col].width = largura
    
    # Criando e aplicando estilo de data
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY', alignment=Alignment(horizontal='center'))
    for row in ws_novo1.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.style = date_style

    # Salvando o arquivo
    nome_arquivo = f"REQ_ABERTAS_STATUS_{datetime.now().strftime('%d%m%Y')}.xlsx"
    caminho_completo = os.path.join(os.path.dirname(file_path), nome_arquivo)
    wb_novo.save(caminho_completo)
    print(f"Arquivo salvo como: {caminho_completo}")

if __name__ == "__main__":
    selecionar_arquivo()
